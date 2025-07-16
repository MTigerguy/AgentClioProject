print("CLIO: This is the current development version running.")
# Streamlined version: Core PDF processing only, merging and DnD removed

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk, simpledialog
from PyPDF2 import PdfReader, PdfWriter 
import fitz      # type: ignore # PyMuPDF
import re
import os
import datetime
import traceback
from openpyxl import Workbook, load_workbook # type: ignore
import csv

# Summary TXT creation
def create_summary_txt(folder, base, borrower, date_of_signing, letter_count, legal_count, other_count, total_count):
    filename = os.path.join(folder, f"{base}_Summary.txt")
    with open(filename, "w", encoding="utf-8") as f:
        f.write("Document Summary\n")
        f.write(f"Date Created: {datetime.date.today().strftime('%Y-%m-%d')}\n")
        f.write(f"Date of Signing: {date_of_signing}\n")
        f.write(f"Borrower(s): {borrower}\n")
        f.write(f"Total Page Count: {total_count}\n")
        f.write(f"  • Letter: {letter_count}\n")
        f.write(f"  • Legal: {legal_count}\n")
        f.write(f"  • Other: {other_count}\n")
    print(f"  Saving summary: {filename}")

# ----------------- Configuration -----------------
BASE_DIR = os.path.expanduser("~/Documents/Agents/AgentClioProject/MAB Law LLC")
LOG_EXCEL = os.path.join(BASE_DIR, "Clio_Log.xlsx")
LOG_CSV = os.path.join(BASE_DIR, "Clio_Log.csv")
ERROR_LOG = os.path.join(BASE_DIR, "Clio_app_error.log")

#show status window
def show_status_window(text, filenames=None):
    win = tk.Toplevel()
    win.title("Processing Results")
    win.geometry("500x400")
    st = scrolledtext.ScrolledText(win, wrap=tk.WORD)
    st.pack(fill=tk.BOTH, expand=True)
    st.insert(tk.END, text)
    # Add output filenames, if provided
    if filenames:
        st.insert(tk.END, "\n\nOutput Files Created:\n")
        for fn in filenames:
            st.insert(tk.END, f"  {fn}\n")
    st.config(state=tk.DISABLED)

# --- Dialog for selecting folder and filename (for fallback/manual naming) ---
def select_folder_and_name():
    # Use a Toplevel dialog (assume a root exists or caller ensures one)
    parent = tk._default_root if tk._default_root else tk.Toplevel()
    folder = filedialog.askdirectory(title="Select Output Folder", parent=parent)
    if not folder:
        if not tk._default_root:
            parent.destroy()
        return None, None
    # Simple prompt for filename
    def on_ok():
        name = name_var.get().strip()
        if not name or not name.replace("_", "").replace("-", "").isalnum() or len(name) > 20:
            messagebox.showerror("Invalid Name", "Enter 1-20 letters/numbers (_ and - allowed).", parent=dialog)
            return
        dialog.result = name
        dialog.destroy()
    dialog = tk.Toplevel(parent)
    dialog.title("Enter Base Filename")
    tk.Label(dialog, text="Base filename (1–20 chars, letters/numbers/_/-):").pack(padx=10, pady=8)
    name_var = tk.StringVar()
    entry = tk.Entry(dialog, textvariable=name_var, width=25)
    entry.pack(padx=10, pady=4)
    entry.focus()
    tk.Button(dialog, text="OK", command=on_ok).pack(pady=10)
    dialog.result = None
    dialog.transient(parent)
    dialog.grab_set()
    parent.wait_window(dialog)
    if not tk._default_root:
        parent.destroy()
    if dialog.result:
        return folder, dialog.result
    return None, None
# —————————————————————————————————————————
# Split + marker insertion + subfolders
def split_and_save_pdfs(path, folder, base, reader):
    """
    Splits the PDF into Letter, Legal, Other, and Full, and saves to disk. Returns a summary string.
    """
    # prepare writers
    letter_writer = PdfWriter()
    legal_writer = PdfWriter()
    other_writer = PdfWriter()

    letter_count = legal_count = other_count = 0
    in_legal_block = False

    for page in reader.pages:
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        tp = get_paper_type(w, h)

        if tp == "Letter":
            letter_writer.add_page(page)
            letter_count += 1
            in_legal_block = False

        elif tp == "Legal":
            legal_writer.add_page(page)
            legal_count += 1
            if not in_legal_block:
                # one marker page in letter stream per legal block
                letter_writer.add_blank_page(width=8.5*72, height=11*72)
                in_legal_block = True

        else:
            other_writer.add_page(page)
            other_count += 1

    # Save Letter
    if letter_count or in_legal_block:
        print(f"  Saving: {os.path.join(folder, f'{base}_Letter.pdf')}")
        with open(os.path.join(folder, f"{base}_Letter.pdf"), "wb") as f:
            letter_writer.write(f)
    # Save Legal
    if legal_count:
        print(f"  Saving: {os.path.join(folder, f'{base}_Legal.pdf')}")
        with open(os.path.join(folder, f"{base}_Legal.pdf"), "wb") as f:
            legal_writer.write(f)
    # Save Other (if any)
    if other_count:
        print(f"  Saving: {os.path.join(folder, f'{base}_Other.pdf')}")
        with open(os.path.join(folder, f"{base}_Other.pdf"), "wb") as f:
            other_writer.write(f)
    # Save Full (always!)
    full_writer = PdfWriter()
    for page in reader.pages:
        full_writer.add_page(page)
    print(f"  Saving: {os.path.join(folder, f'{base}_Full.pdf')}")
    with open(os.path.join(folder, f"{base}_Full.pdf"), "wb") as f:
        full_writer.write(f)

    # Prompt for date of signing
    date_of_signing = simpledialog.askstring("Date of Signing", "Enter date of signing (YYYY-MM-DD):")
    # Derive borrower name from base
    borrower_name = base.split("_")[0]
    # Total count
    total_count = len(reader.pages)
    # Call summary TXT function
    create_summary_txt(folder, base, borrower_name, date_of_signing, letter_count, legal_count, other_count, total_count)

    return f"{os.path.basename(path)} → Letter:{letter_count}, Legal:{legal_count}, Other:{other_count}\n"


def process_pdfs_individually_with_filelist(file_paths):
    if not file_paths:
        return

    log_txt = ""
    output_filenames = []
    for path in file_paths:
        name = os.path.basename(path)
        print(f"Processing: {path}")
        try:
            folder, base = extract_base_filename(path)
            print(f"  Extracted folder: {folder}")
            print(f"  Extracted base: {base}")
            if not folder or not base:
                skip_reason = f"{os.path.basename(path)} → SKIPPED (No name/folder)"
                print(f"  {skip_reason}")
                log_txt += skip_reason + "\n"
                continue
            os.makedirs(folder, exist_ok=True)
            reader = PdfReader(path)
            # Call split_and_save_pdfs and collect output filenames
            summary = split_and_save_pdfs(path, folder, base, reader)
            log_txt += summary
            # Add all output files that are created
            date_str = datetime.date.today().strftime("%m%d%Y")
            # These files are always named as base + _Letter, _Legal, _Other, _Full.pdf
            # But only created if count > 0 (except Full, always)
            # So check which files exist and add to list
            for typ in ["Letter", "Legal", "Other", "Full"]:
                out_path = os.path.join(folder, f"{base}_{typ}.pdf")
                if os.path.exists(out_path):
                    output_filenames.append(out_path)
        except Exception as e:
            print(f"  ERROR processing {path}: {e}")
            log_txt += f"{os.path.basename(path)} → ERROR: {str(e)}\n"
            continue

    # Show status window with all results and output files
    show_status_window(log_txt, filenames=output_filenames)
    log_action("Process", [os.path.basename(f) for f in file_paths], log_txt)
    print("Batch processing complete. See above for details.")

#intake Screen 
def show_intake_window():
    intake_win = tk.Tk()
    intake_win.title("Clio Document Intake")
    intake_win.geometry("900x350")

    selected_files = []

    def select_files():
        files = filedialog.askopenfilenames(
            title="Select PDF Files", filetypes=[("PDF files", "*.pdf")]
        )
        if files:
            selected_files.clear()
            selected_files.extend(files)
            file_listbox.delete(0, tk.END)
            for f in files:
                file_listbox.insert(tk.END, f)

    def process_files():
        if not selected_files:
            messagebox.showerror("No Files", "Please select PDFs first.")
            return
        process_pdfs_individually_with_filelist(selected_files)

    def refresh():
        selected_files.clear()
        file_listbox.delete(0, tk.END)

    def exit_app():
        intake_win.destroy()

    file_listbox = tk.Listbox(intake_win, width=100, height=10)
    file_listbox.pack(pady=10)

    tk.Button(intake_win, text="Select PDFs", command=select_files).pack()

    btn_frame = tk.Frame(intake_win)
    btn_frame.pack(pady=20)

    tk.Button(btn_frame, text="Process", width=15, command=process_files).grid(row=0, column=0, padx=5)
    tk.Button(btn_frame, text="Refresh", width=15, command=refresh).grid(row=0, column=1, padx=5)
    tk.Button(btn_frame, text="Exit", width=15, command=exit_app).grid(row=0, column=2, padx=5)

    intake_win.mainloop()
# --- Manual entry popup for name if not found automatically ---
def manual_name_prompt():
    class NamePrompt(simpledialog.Dialog):
        def body(self, master):
            tk.Label(master, text="Last Name (max 10 letters):").grid(row=0)
            tk.Label(master, text="First Initial:").grid(row=1)
            self.last_var = tk.StringVar()
            self.init_var = tk.StringVar()
            self.last_entry = tk.Entry(master, textvariable=self.last_var)
            self.init_entry = tk.Entry(master, textvariable=self.init_var, width=4)
            self.last_entry.grid(row=0, column=1)
            self.init_entry.grid(row=1, column=1)
            return self.last_entry

        def validate(self):
            last = self.last_var.get().strip()
            initial = self.init_var.get().strip().upper()
            print(f"DEBUG: last='{last}', initial='{initial}'")
            # Last: required, 1-10 letters
            if not last or len(last) > 10 or not last.isalpha():
                messagebox.showerror("Error", "Last name must be 1–10 letters.", parent=self)
                return False
            # First initial: required, 1-2 letters only
            if not initial or not (1 <= len(initial) <= 2) or not initial.isalpha():
                messagebox.showerror("Error", "First initial must be 1 or 2 letters.", parent=self)
                return False
            return True

        def apply(self):
            self.result = (self.last_var.get().strip().capitalize(), self.init_var.get().strip().upper())

    parent = tk._default_root if tk._default_root else tk.Toplevel()
    prompt = NamePrompt(parent, "Filename: Cannot find name. Please provide last name (max 10 letters) and first initial.")
    if not tk._default_root:
        parent.destroy()
    if hasattr(prompt, "result") and prompt.result and prompt.result[0] and prompt.result[1]:
        return f"{prompt.result[0]}_{prompt.result[1]}"
    else:
        return None

# —————————————————————————————————————————
# Helper: extract base filename from PDF content
def extract_base_filename(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""

    # --- Pages to skip if they contain agent/instruction keywords ---
    instruction_filters = [
        "attention closing agent",
        "closing agent instruction",
        "instructions to the closing agent",
        "attention signing agent",
        "notary checklist",
        "special instructions",
        "agent acknowledgement",
    ]

    # --- Expanded list of words/roles that should never be used as borrower names ---
    not_borrower_filters = [
        "lender", "lenders", "lender name", "lender rep", "lender representative",
        "mortgage company", "mortgage companies", "bank", "banks", "servicer", "servicers",
        "broker", "brokers", "realty", "real estate", "real estate agent", "real estate agents",
        "title", "title agent", "title agents", "title rep", "title reps", "title company", "title companies",
        "escrow officer", "escrow officers", "settlement agent", "settlement agents", "notary", "notaries",
        "witness", "witnesses", "signature", "signatures", "loan officer", "loan officers", "processor", "processors",
        "signing agent", "signing agents", "closing agent", "closing agents", "underwriter", "underwriters",
        "attorney", "attorneys", "law firm", "law firms", "company", "companies", "inc", "inc.", "llc", "llc.", "corp",
        "corporation", "corporations", "co.", "co", "llp", "llp.", "pllc", "pllc.", "plc", "plc.", "pa", "p.a.", "pc", "p.c.",
        "associates", "associate", "group", "groups", "firm", "firms", "office", "offices", "department", "departments",
        "division", "divisions", "section", "sections", "admin", "administrator", "administrators", "administer",
        "president", "presidents", "vice president", "secretary", "secretaries", "manager", "managers", "management",
        "director", "directors", "officer", "officers", "official", "officials", "contact", "contacts", "employee", "employees",
        "staff", "team", "teams", "counsel", "adviser", "advisor", "consultant", "consultants", "independent contractor",
        "independent contractors", "contractor", "contractors", "organizer", "organizers", "participant", "participants",
        "benefactor", "benefactors", "grantor", "grantors", "grantee", "grantees", "remitter", "remitters",
        "payee", "payees", "payor", "payors", "mortgagor", "mortgagors", "mortgagee", "mortgagees",
        "trust", "trusts", "trustee", "trustees", "foundation", "foundations", "estate", "estates", "heir", "heirs",
        "beneficiary", "beneficiaries", "power of attorney", "poa", "personal representative", "personal representatives",
        "successor", "successors", "authorized", "representative", "representatives", "agent", "agents", "approved",
        "accept", "accepted", "seller", "sellers", "buyer", "buyers", "borrower", "borrowers", "co-borrower", "co-borrowers",
        "joint tenant", "joint tenants", "spouse", "spouses", "partner", "partners", "appointee", "appointees",
        "recipient", "recipients", "customer", "customers", "client", "clients", "occupant", "occupants",
        "landlord", "landlords", "tenant", "tenants", "lessee", "lessees", "lessor", "lessors", "guarantor", "guarantors",
        "north charleston", "charleston", "clearedge", "deceased", "account", "accounts", "section", "sections",
        "division", "divisions", "administer", "controller", "supervisor", "supervisors", "applicant", "applicants"
    ]
    generic_labels = {"borrower", "owner", "seller", "buyer", "applicant", "customer", "client"}

    # --- Concatenate non-instruction pages ---
    for i in range(min(20, doc.page_count)):
        page_text = doc[i].get_text()
        page_start = page_text[:300].lower()
        # Smarter skip: If a page is instructions, but has borrower info, keep it
        if any(filter_text in page_start for filter_text in instruction_filters):
            if ("homeowner name" in page_text.lower()) or ("borrower" in page_text.lower()) or ("applicant" in page_text.lower()):
                text += page_text
                continue
            continue
        text += page_text

    # --- Helper function: clean and validate candidate name ---
    def is_valid_name(candidate):
        candidate_lower = candidate.lower().strip()
        # Reject if it's a generic label exactly
        if candidate_lower in generic_labels:
            return False
        # Reject if matches any filter term (substring)
        if any(nb in candidate_lower for nb in not_borrower_filters):
            return False
        # Reject if contains digits or odd characters
        if re.search(r"[^a-zA-Z ,.'&-]", candidate):
            return False
        # Reject if too short
        if len(candidate) < 3:
            return False
        return True

    # --- Main extraction: try known labels with regex ---
    name = None
    patterns = [
        r"Borrower Information\s*[:\n]+([A-Za-z ,.'&-]+)",
        r"Borrower\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Borrower:\s*([A-Za-z ,.'&-]+)",
        r"Homeowner Name\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Owner\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Property Owner\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Seller\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Buyer\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Applicant\(s\):\s*([A-Za-z ,.'&-]+)",
        r"Client:\s*([A-Za-z ,.'&-]+)",
        r"Customer:\s*([A-Za-z ,.'&-]+)",
    ]

    # Try pattern-based extraction
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            candidate = m.group(1).split(' and ')[0].split(',')[0].strip()
            candidate_lower = candidate.lower().strip()
            if candidate_lower in generic_labels or candidate_lower in not_borrower_filters:
                continue  # skip generic/filtered names
            if is_valid_name(candidate):
                name = candidate
                break

    # If not found, look for sequences of capitalized words (name heuristics)
    if not name:
        cap_name_matches = re.findall(r"\b([A-Z][a-z]+(?: [A-Z][a-z]+)+)\b", text)
        for cand in cap_name_matches:
            cand_lower = cand.lower().strip()
            if cand_lower in generic_labels or cand_lower in not_borrower_filters:
                continue
            if is_valid_name(cand):
                name = cand
                break

    # Fallback: try all-caps 2-word patterns
    if not name:
        acaps = re.findall(r"\n([A-Z]{2,} [A-Z]{2,})\n", text)
        for acap in acaps:
            acap_title = acap.title()
            acap_lower = acap_title.lower().strip()
            if acap_lower in generic_labels or acap_lower in not_borrower_filters:
                continue
            if is_valid_name(acap):
                name = acap_title
                break

    # If no real name is found, always show the manual name prompt as fallback
    if not name or name.lower().strip() in generic_labels or name.lower().strip() in not_borrower_filters or name.lower() == "unknown":
        base_name = manual_name_prompt()
        if base_name:
            folder = BASE_DIR
            date_str = datetime.date.today().strftime("%m%d%Y")
            folder_path = os.path.join(folder, datetime.date.today().isoformat(), base_name)
            return folder_path, f"{base_name}_{date_str}"
        else:
            return None, None
    else:
        name_parts = name.split()
        if len(name_parts) >= 2:
            last = name_parts[-1][:10]  # Max 10 letters
            first_initial = name_parts[0][0]
        else:
            last = name_parts[0][:10]
            first_initial = name_parts[0][0] if name_parts else "X"
        borrower_part = f"{last}_{first_initial}"
        date_str = datetime.date.today().strftime("%m%d%Y")

        today_str = datetime.date.today().isoformat()
        folder = os.path.join(BASE_DIR, today_str, borrower_part)
        return folder, f"{borrower_part}_{date_str}"

# —————————————————————————————————————————
# Logging: Excel + CSV
def log_action(action_type, filenames, output_path):
    os.makedirs(BASE_DIR, exist_ok=True)

    # Excel log
    if not os.path.exists(LOG_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Date", "Action", "Filenames", "Output Path"])
    else:
        wb = load_workbook(LOG_EXCEL)
        ws = wb["Log"]

    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    files_str = ", ".join(filenames)
    ws.append([ts, action_type, files_str, output_path])
    wb.save(LOG_EXCEL)

    # CSV log
    new_file = not os.path.exists(LOG_CSV)
    with open(LOG_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if new_file:
            writer.writerow(["Date", "Action", "Filenames", "Output Path"])
        writer.writerow([ts, action_type, files_str, output_path])

# —————————————————————————————————————————
# Helper to classify paper type
def get_paper_type(w_pts, h_pts):
    w_in, h_in = sorted([w_pts/72, h_pts/72])
    if abs(h_in-11) < 0.2 and abs(w_in-8.5) < 0.2:
        return "Letter"
    if abs(h_in-14) < 0.2 and abs(w_in-8.5) < 0.2:
        return "Legal"
    return "Other"

# Show scrollable window with all selected files
def show_selected_files_window(file_paths, action="Files Selected"):
    win = tk.Toplevel()
    win.title(action)
    win.geometry("500x400")
    st = scrolledtext.ScrolledText(win, wrap=tk.WORD)
    st.pack(fill=tk.BOTH, expand=True)
    st.insert(tk.END, "\n".join(file_paths))
    st.config(state=tk.DISABLED)

# Create a numbered marker PDF page using PyMuPDF
def create_marker_pdf(page_number, width=8.5*72, height=11*72):
    doc = fitz.open()
    page = doc.new_page(width=width, height=height)
    text = f"Marker Page {page_number}"
    rect = fitz.Rect(72, height/2-20, width-72, height/2+20)
    page.insert_textbox(rect, text, fontsize=36, color=(0, 0, 0), align=1)
    # Save to memory, return raw bytes
    pdf_bytes = doc.write()
    doc.close()
    return pdf_bytes

# Log viewer
def view_log():
    if not os.path.exists(LOG_EXCEL):
        messagebox.showerror("Log Missing", "No log file found.")
        return

    def do_filter():
        date_f = date_entry.get().strip()
        act_f = action_var.get()
        wb = load_workbook(LOG_EXCEL)
        ws = wb["Log"]
        results.delete(1.0, tk.END)
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, action, files, outp = row
            if (not date_f or date_f in date) and (act_f == "All" or act_f == action):
                results.insert(tk.END, f"{date} | {action}\nFiles: {files}\n→ {outp}\n\n")

    top = tk.Toplevel()
    top.title("Clio Log Viewer")
    top.geometry("600x500")
    frm = tk.Frame(top); frm.pack(pady=10)
    tk.Label(frm, text="Date (YYYY-MM-DD):").grid(row=0, column=0)
    date_entry = tk.Entry(frm, width=15); date_entry.grid(row=0, column=1, padx=5)
    tk.Label(frm, text="Action:").grid(row=0, column=2)
    action_var = tk.StringVar(value="All")
    ttk.Combobox(frm, textvariable=action_var, values=["All", "Merge", "Process"],
                 state="readonly", width=10).grid(row=0, column=3, padx=5)
    tk.Button(frm, text="Filter", command=do_filter).grid(row=0, column=4, padx=10)
    results = scrolledtext.ScrolledText(top, wrap=tk.WORD)
    results.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)


def process_pdfs_individually():
    file_paths = filedialog.askopenfilenames(
        title="Select PDF Files to Process", filetypes=[("PDF files","*.pdf")])
    if not file_paths:
        return
    process_pdfs_individually_with_filelist(file_paths)


# —————————————————————————————————————————
# GUI
def main():
    show_intake_window()

if __name__ == "__main__":
    try:
        main()
    except Exception:
        # Log full traceback to a file
        with open(ERROR_LOG, "a") as f:
            f.write(traceback.format_exc() + "\n")
        # Show a popup (use Toplevel to avoid extra root window)
        parent = tk._default_root if tk._default_root else tk.Toplevel()
        messagebox.showerror(
            "Application Error",
            f"An unexpected error occurred.\n\n"
            f"The full error has been logged to:\n{ERROR_LOG}",
            parent=parent
        )
        if not tk._default_root:
            parent.destroy()
